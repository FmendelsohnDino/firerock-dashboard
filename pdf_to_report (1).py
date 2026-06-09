#!/usr/bin/env python3
"""
FireRock  PDF -> sales_report.xlsx  converter
=============================================
Turns the two daily printouts (Dashboard.pdf + DailyOrderSummary.pdf) into a
schema-correct sales_report.xlsx that extract_data.py reads with NO changes.

Why this exists
---------------
The daily printout carries everything that moves day-to-day: sales by rep,
the product-line breakdown, revenue MTD/YTD, and backlog. The ONE thing it does
not carry is the Jan->May monthly revenue trend (a separate pivot). Those are
closed months that don't change, so we splice them in from a stored "trend base"
workbook (the last full Excel master export).

Result: every dashboard tab is the SAME date as the printout. No cross-tab
discrepancy. The only thing that lags is the monthly-trend chart's history,
which is historical by nature.

Usage
-----
    python pdf_to_report.py Dashboard.pdf DailyOrderSummary.pdf \
        --base trend_base.xlsx --out sales_report.xlsx

If --base is omitted the REVENUE/trend tab is written empty (chart blank, but
everything else is current). Refresh trend_base.xlsx from a real master export
about once a month so the trend stays reasonably fresh.

Numbers are read by PIXEL POSITION, not text order, because the PDF splits
figures across stray spaces ("5 8,264"). Fragments that physically touch on the
page are re-joined; real column gaps (10px+) are kept apart. Every total is then
cross-footed against its parts and any mismatch is reported loudly so a
mis-parse can never ship silently.
"""
import sys, os, re, argparse
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

GAP = 3.0          # px; below this, two word fragments are the same number
ROW_BIN = 2        # px; rows within this top-coordinate band are one line

REPS = ['ERIC BLASKOWSKI','KEVIN KEITH','THORNTON COLE','MICHAEL KUHLMAN','HAMP BRILEY',
        'GRAYSON MOOREHEAD','BEAU BOUTHILLIER','SUZANNE HORST','TIFFANY HASKEKER']
REP_TOTAL = 'TOTAL FIREROCK SALES TEAM'
PRODUCTS = ['FIREPLACE/PP','PAVER/PP','ROOFING','FLOORING','INSTALLATION','SWD',
            'OTHER','TOTAL','FREIGHT','TOTAL W/ FREIGHT']


# ----------------------------------------------------------------------------- parsing
def _merge(words, gap=GAP):
    words = sorted(words, key=lambda w: w['x0'])
    out = []
    for w in words:
        if out and (w['x0'] - out[-1]['x1']) < gap:
            out[-1]['text'] += w['text']; out[-1]['x1'] = w['x1']
        else:
            out.append({'text': w['text'], 'x0': w['x0'], 'x1': w['x1']})
    return out


def _num(t):
    t = t.strip()
    if t in ('-', '', '\u2013'): return 0
    if t.endswith('%'):
        try: return float(t[:-1]) / 100
        except: return 0.0
    try: return float(t.replace(',', ''))
    except: return t


def _is_num(t):
    return t == '-' or bool(re.match(r'^-?[\d,]+%?$', t))


def parse_dashboard(path):
    """Return ({despaced_label: [values...]}, date_text, meta) for the dashboard."""
    with pdfplumber.open(path) as pdf:
        words = pdf.pages[0].extract_words()
        text_lines = pdf.pages[0].extract_text().split('\n')
    date_text = text_lines[2].strip()

    # header meta: "# Days In 6 27% # Days In 112 44%" and working-days line
    meta = {'workDaysMonth': 22, 'workDaysYear': 254,
            'daysIn': 6, 'pctMonth': 0.0, 'ytdDays': 112, 'pctYear': 0.0}
    for ln in text_lines:
        m = re.search(r'Working Days in Current Month\s+(\d+).*Current Year\s+(\d+)', ln)
        if m:
            meta['workDaysMonth'] = int(m.group(1)); meta['workDaysYear'] = int(m.group(2))
        m = re.search(r'#\s*Days In\s+(\d+)\s+(\d+)%.*#\s*Days In\s+(\d+)\s+(\d+)%', ln)
        if m:
            meta['daysIn'] = int(m.group(1)); meta['pctMonth'] = int(m.group(2)) / 100
            meta['ytdDays'] = int(m.group(3)); meta['pctYear'] = int(m.group(4)) / 100

    rows = {}
    for w in words:
        rows.setdefault(round(w['top'] / ROW_BIN) * ROW_BIN, []).append(w)
    parsed = {}
    for key in sorted(rows):
        toks = [m['text'] for m in _merge(rows[key])]
        label, vals, in_vals = [], [], False
        for t in toks:
            if not _is_num(t) and not in_vals:
                label.append(t)
            else:
                in_vals = True; vals.append(_num(t))
        lab = ''.join(label).replace(' ', '').upper()
        if lab:
            parsed[lab] = vals
    return parsed, date_text, meta


def _key(name):
    return name.replace(' ', '').replace('/', '/').upper()


def parse_orders(path):
    """Daily Order Summary -> list of dicts. Uses plain text lines (spaces intact);
    skips per-rep subtotals and the grand total."""
    out = []
    with pdfplumber.open(path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    current_rep = ''
    for line in text.split('\n'):
        line = line.strip()
        if not line or 'Total' in line or 'Grand' in line:
            # a "<NAME> Total" line still updates nothing; skip
            continue
        m = re.search(r'\b(FRKO\d+(?:-P)?)\b', line)
        if m:
            so = m.group(1)
            pre = line[:m.start()].strip()
            rest = line[m.end():].strip()
            # trailing amount = last token
            parts = rest.rsplit(' ', 1)
            if len(parts) == 2 and re.match(r'^[\d,]+$', parts[1].replace(',', '')):
                customer, amt = parts[0].strip(), _num(parts[1])
            else:
                customer, amt = rest, 0
            # a rep name may lead this line (caps) before the SO
            if pre and pre.upper() == pre:
                current_rep = pre
            out.append({'rep': current_rep, 'so': so,
                        'customer': customer, 'amount': amt if isinstance(amt,(int,float)) else 0})
        else:
            # standalone rep header line (all caps, no SO, no amount)
            if line.upper() == line and not re.search(r'\d', line):
                current_rep = line
    return out


# ----------------------------------------------------------------------------- validation
def validate(dash, reps, total, prods, errors):
    # Cross-check our summed team total against the printed TOTAL row (first value = MTD).
    printed = dash.get(_key(REP_TOTAL), [])
    if printed:
        pm = printed[0]
        if abs(pm - total['mtd']) > max(1.0, 0.005 * max(pm, 1)):
            errors.append(f"Summed rep MTD {total['mtd']:,.0f} != printed team total {pm:,.0f}")
    # Product TOTAL row should foot to the sales team total.
    pt = prods.get('TOTAL', {}).get('mtd', 0)
    # (product TOTAL includes corp/house lines, so allow it to be >= team; just sanity-check non-zero)
    if total['mtd'] and pt and pt < total['mtd'] * 0.5:
        errors.append(f"Product TOTAL MTD {pt:,.0f} implausibly low vs team {total['mtd']:,.0f}")
    if not (0 <= (total['pctQ'] or 0) <= 2):
        errors.append(f"team %quota out of range: {total['pctQ']}")
    # Cross-check derived TOTAL W/ FREIGHT revenue against numbers printed on that row.
    wf = prods.get('_wfreight', {})
    raw = set(round(x) for x in dash.get('TOTALW/FREIGHT', []) if isinstance(x, (int, float)))
    for label, val in (('revMTD', wf.get('revMTD')), ('revYTD', wf.get('revYTD'))):
        if val and round(val) not in raw:
            errors.append(f"derived w/freight {label} {val:,.0f} not found on printed TOTAL W/FREIGHT row")
    return errors


# ----------------------------------------------------------------------------- writing
def build_workbook(dash, date_text, meta, orders, base_path, out_path):
    def gv(label, idx, default=0):
        v = dash.get(_key(label), [])
        return v[idx] if idx < len(v) else default

    wb = Workbook()
    ws = wb.active; ws.title = "DASHBOARD"
    sub = PatternFill("solid", fgColor="2E75B6")
    tot = PatternFill("solid", fgColor="BDD7EE")
    alt = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style='thin', color='AAAAAA'); bd = Border(thin, thin, thin, thin)

    def put(r, c, v, *, bold=False, fill=None, fmt=None, align="right", color="000000", size=10):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="Arial", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        cell.border = bd
        return cell

    try:
        dt = datetime.strptime(date_text, "%m/%d/%Y")
    except Exception:
        dt = datetime.now()
    ws.merge_cells("A1:S1")
    put(1,1,"FireRock Building Materials \u2014 Daily Sales & Revenue Report", bold=True,
        fill=PatternFill("solid", fgColor="1F3864"), color="FFFFFF", align="center", size=13)
    put(3,1,"Report Date", bold=True, align="left"); put(3,2,dt, fmt='m/d/yyyy', align="left")

    put(5,3,"# Working Days in Current Month", align="left"); put(5,4, meta['workDaysMonth'])
    put(5,10,"# Working Days in Current Year", align="left"); put(5,11, meta['workDaysYear'])
    put(6,3,"# Days In", align="left"); put(6,4, meta['daysIn']); put(6,5, meta['pctMonth'], fmt='0%')
    put(6,10,"# Days In", align="left"); put(6,11, meta['ytdDays']); put(6,12, meta['pctYear'], fmt='0%')

    # ------- rep rows 14-23  (rep layout is STABLE; idx 0 mtd,1 py,2 quota,3 %q,4 fore,
    #                          5 fore%,6 backlog,8 ytd26,9 ytd25,13 annQ)
    col_hdr = {2:"Sales Rep",3:"MTD",4:"PY",5:"Month Quota",6:"% Quota",7:"Forecast",
               8:"Fore %Q",12:"Backlog",14:"YTD 2026",15:"YTD 2025",19:"Annual Quota"}
    for c,l in col_hdr.items():
        put(13,c,l, bold=True, fill=sub, color="FFFFFF", align="center", size=9)
    reps = {}
    for i, name in enumerate(REPS):
        v = dash.get(_key(name), [])
        g = lambda k, d=0: v[k] if k < len(v) else d
        reps[name] = {'mtd':g(0),'py':g(1),'quota':g(2),'pctQ':g(3),'fore':g(4),
                      'forePct':g(5),'backlog':g(6),'ytd':g(8),'ytdPY':g(9),'annQ':g(13)}
        r = 14 + i; f = alt if i % 2 == 0 else None
        put(r,2,name, bold=True, align="left", fill=f)
        put(r,3,g(0), fmt='#,##0', fill=f); put(r,4,g(1), fmt='#,##0', fill=f)
        put(r,5,g(2), fmt='#,##0', fill=f); put(r,6,g(3), fmt='0%', fill=f)
        put(r,7,g(4), fmt='#,##0', fill=f); put(r,8,g(5), fmt='0%', fill=f)
        put(r,12,g(6), fmt='#,##0', fill=f); put(r,14,g(8), fmt='#,##0', fill=f)
        put(r,15,g(9), fmt='#,##0', fill=f); put(r,19,g(13), fmt='#,##0', fill=f)

    # ------- team total: SUM the rep rows (robust; the printed total row drops a
    #         column and would shift indices). Printed total kept only for cross-check.
    sk = lambda k: sum(reps[r][k] for r in REPS)
    total = {'mtd':sk('mtd'),'py':sk('py'),'quota':sk('quota'),'fore':sk('fore'),
             'backlog':sk('backlog'),'ytd':sk('ytd'),'ytdPY':sk('ytdPY'),'annQ':sk('annQ')}
    total['pctQ'] = total['mtd']/total['quota'] if total['quota'] else 0
    total['forePct'] = total['fore']/total['quota'] if total['quota'] else 0
    for rr in (23, 30):
        put(rr,2,REP_TOTAL, bold=True, align="left", fill=tot)
        for c,k,fmt in [(3,'mtd','#,##0'),(4,'py','#,##0'),(5,'quota','#,##0'),(6,'pctQ','0%'),
                        (7,'fore','#,##0'),(8,'forePct','0%'),(12,'backlog','#,##0'),
                        (14,'ytd','#,##0'),(15,'ytdPY','#,##0'),(19,'annQ','#,##0')]:
            put(rr,c,total[k], bold=True, fmt=fmt, fill=tot)

    # ------- product rows 36-45. Product LINE rows (FIREPLACE..OTHER) have a stable
    #         layout: idx 0 mtd, 2 quota, 3 %q, 6 revMTD, 7 revBgtMTD, 8 revPct,
    #         9 backlog, 10 revYTD, 11 revBgtYTD. The TOTAL / FREIGHT / TOTAL W/FREIGHT
    #         rows are irregular (dropped %-cells, "-" in quota), so we DERIVE them by
    #         summation instead of trusting their column positions.
    LINES = ['FIREPLACE/PP','PAVER/PP','ROOFING','FLOORING','INSTALLATION','SWD','OTHER']
    prods = {}
    line_sum = {'mtd':0,'revMTD':0,'revBgtMTD':0,'revYTD':0,'revBgtYTD':0,'backlog':0}
    for i, name in enumerate(LINES):
        v = dash.get(_key(name), [])
        g = lambda k, d=0: v[k] if k < len(v) else d
        rec = {'mtd':g(0),'quota':g(2),'pctQ':g(3),'revMTD':g(6),'revBgtMTD':g(7),
               'revPct':g(8),'backlog':g(9),'revYTD':g(10),'revBgtYTD':g(11)}
        prods[name] = {'mtd':rec['mtd'],'revMTD':rec['revMTD'],'revYTD':rec['revYTD']}
        for k in line_sum: line_sum[k] += rec.get(k, 0)
        r = 36 + i
        put(r,2,name, bold=True, align="left")
        put(r,3,rec['mtd'], fmt='#,##0'); put(r,5,rec['quota'], fmt='#,##0')
        put(r,6,rec['pctQ'], fmt='0%'); put(r,10,rec['revMTD'], fmt='#,##0')
        put(r,11,rec['revBgtMTD'], fmt='#,##0'); put(r,12,rec['revPct'], fmt='0%')
        put(r,14,rec['backlog'], fmt='#,##0'); put(r,16,rec['revYTD'], fmt='#,##0')

    # FREIGHT row: no SO data, so positional index is unreliable; take its nonzero
    # numeric run = [revMTD, revBgtMTD, %, revYTD, revBgtYTD, %].
    fv = [x for x in dash.get('FREIGHT', []) if isinstance(x, (int, float)) and x != 0]
    freight = {'revMTD': fv[0] if len(fv) > 0 else 0,
               'revBgtMTD': fv[1] if len(fv) > 1 else 0,
               'revYTD': fv[3] if len(fv) > 3 else 0,
               'revBgtYTD': fv[4] if len(fv) > 4 else 0}

    # TOTAL (row 43) and TOTAL W/ FREIGHT (row 45) by summation
    total_prod = dict(line_sum)
    wfreight = {'revMTD': total_prod['revMTD'] + freight['revMTD'],
                'revBgtMTD': total_prod['revBgtMTD'] + freight['revBgtMTD'],
                'revYTD': total_prod['revYTD'] + freight['revYTD'],
                'revBgtYTD': total_prod['revBgtYTD'] + freight['revBgtYTD'],
                'mtd': total_prod['mtd']}
    wfreight['revPctMTD'] = wfreight['revMTD']/wfreight['revBgtMTD'] if wfreight['revBgtMTD'] else 0
    wfreight['revPctYTD'] = wfreight['revYTD']/wfreight['revBgtYTD'] if wfreight['revBgtYTD'] else 0

    put(43,2,"TOTAL", bold=True, align="left")
    put(43,3,total_prod['mtd'], fmt='#,##0'); put(43,10,total_prod['revMTD'], fmt='#,##0')
    put(43,14,total_prod['backlog'], fmt='#,##0'); put(43,16,total_prod['revYTD'], fmt='#,##0')
    put(44,2,"FREIGHT", bold=True, align="left")
    put(44,10,freight['revMTD'], fmt='#,##0'); put(44,16,freight['revYTD'], fmt='#,##0')
    # row 45 = headline-revenue source the script reads (c10,c11,c12,c16,c17,c18)
    put(45,2,"TOTAL W/ FREIGHT", bold=True, align="left")
    put(45,3,wfreight['mtd'], fmt='#,##0')
    put(45,10,wfreight['revMTD'], fmt='#,##0'); put(45,11,wfreight['revBgtMTD'], fmt='#,##0')
    put(45,12,wfreight['revPctMTD'], fmt='0%'); put(45,16,wfreight['revYTD'], fmt='#,##0')
    put(45,17,wfreight['revBgtYTD'], fmt='#,##0'); put(45,18,wfreight['revPctYTD'], fmt='0%')
    prods['TOTAL'] = {'mtd':total_prod['mtd'],'revMTD':total_prod['revMTD'],'revYTD':total_prod['revYTD']}
    prods['_wfreight'] = wfreight

    ws.column_dimensions['A'].width = 4; ws.column_dimensions['B'].width = 22
    for c in range(3,20): ws.column_dimensions[get_column_letter(c)].width = 11

    # ------- REVENUE trend tab (spliced from base; else empty scaffold)
    wr = wb.create_sheet("REVENUE")
    for c,h in enumerate(["SALES REP","Jan","Feb","Mar","Apr","May","Grand Total"], 1):
        cell = wr.cell(4,c,h); cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = sub; cell.alignment = Alignment(horizontal="center")
    trend_rows = []
    if base_path and os.path.exists(base_path):
        try:
            bwb = load_workbook(base_path, data_only=True)
            if 'REVENUE' in bwb.sheetnames:
                br = bwb['REVENUE']
                for r in range(5, 21):
                    nm = br.cell(r,1).value
                    if nm:
                        trend_rows.append([nm] + [br.cell(r,c).value for c in range(2,8)])
        except Exception as e:
            print(f"  (warning: could not read trend base: {e})")
    for i, row in enumerate(trend_rows):
        for c, val in enumerate(row, 1):
            cell = wr.cell(5 + i, c, val)
            cell.font = Font(name="Arial", size=10)
            if c > 1: cell.number_format = '#,##0'
    wr.column_dimensions['A'].width = 22

    # ------- ORDERS tab
    wo = wb.create_sheet("ORDERS")
    for c,h in enumerate(["Sales Rep","SO #","Customer","Amount","Note"], 1):
        cell = wo.cell(1,c,h); cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = sub; cell.alignment = Alignment(horizontal="center")
    for i, o in enumerate(orders):
        r = 2 + i
        wo.cell(r,1,o['rep']).font = Font(name="Arial", size=10)
        wo.cell(r,2,o['so']).font = Font(name="Arial", size=10)
        wo.cell(r,3,o['customer']).font = Font(name="Arial", size=10)
        ac = wo.cell(r,4,o['amount']); ac.number_format = '$#,##0'; ac.font = Font(name="Arial", size=10)
    wo.column_dimensions['A'].width = 22; wo.column_dimensions['B'].width = 16
    wo.column_dimensions['C'].width = 38; wo.column_dimensions['D'].width = 12

    wb.save(out_path)
    return reps, total, prods


def _classify_pdfs(folder):
    """Given a folder of PDFs, return (dashboard_path, orders_path) by content,
    so it doesn't matter what the files are named inside the daily ZIP."""
    dash_pdf = orders_pdf = None
    for fn in sorted(os.listdir(folder)):
        if not fn.lower().endswith('.pdf'):
            continue
        path = os.path.join(folder, fn)
        try:
            with pdfplumber.open(path) as pdf:
                head = (pdf.pages[0].extract_text() or "")[:400].upper()
        except Exception:
            continue
        if 'SALES ORDERS BY DAY' in head:
            orders_pdf = path
        elif 'DAILY SALES' in head or 'SALES ORDERS - CURRENT MONTH' in head:
            dash_pdf = path
    return dash_pdf, orders_pdf


def _unzip(zip_path):
    import zipfile, tempfile
    tmp = tempfile.mkdtemp(prefix="firerock_")
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(tmp)
    # PDFs may sit in a subfolder; flatten by searching
    for root, _, files in os.walk(tmp):
        if any(f.lower().endswith('.pdf') for f in files):
            return root
    return tmp


# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dashboard_pdf", nargs="?", help="Dashboard.pdf (omit if using --zip)")
    ap.add_argument("orders_pdf", nargs="?", help="DailyOrderSummary.pdf (omit if using --zip)")
    ap.add_argument("--zip", dest="zip_path", default=None,
                    help="daily ZIP containing both PDFs (auto-detected by content)")
    ap.add_argument("--base", default=None, help="master .xlsx to splice monthly trend from")
    ap.add_argument("--out", default="sales_report.xlsx")
    args = ap.parse_args()

    if args.zip_path:
        folder = _unzip(args.zip_path)
        dpdf, opdf = _classify_pdfs(folder)
        if not dpdf or not opdf:
            print(f"ERROR: could not find both PDFs in {args.zip_path} "
                  f"(dashboard={dpdf}, orders={opdf})")
            sys.exit(1)
    else:
        dpdf, opdf = args.dashboard_pdf, args.orders_pdf
        if not dpdf or not opdf:
            print("ERROR: provide two PDFs or --zip"); sys.exit(1)

    dash, date_text, meta = parse_dashboard(dpdf)
    orders = parse_orders(opdf)
    reps, total, prods = build_workbook(dash, date_text, meta, orders, args.base, args.out)

    errors = []
    validate(dash, reps, total, prods, errors)
    print(f"Report date : {date_text}")
    print(f"Reps parsed : {len(reps)}   Orders parsed : {len(orders)}")
    print(f"Team MTD    : ${total['mtd']:,.0f}   YTD ${total['ytd']:,.0f}   Annual Q ${total['annQ']:,.0f}")
    fp = prods.get('FIREPLACE/PP', {})
    print(f"FIREPLACE/PP: MTD ${fp.get('mtd',0):,.0f}  revMTD ${fp.get('revMTD',0):,.0f}  revYTD ${fp.get('revYTD',0):,.0f}")
    if errors:
        print("\n*** VALIDATION WARNINGS — review before uploading ***")
        for e in errors: print("  -", e)
        sys.exit(2)
    print(f"\nOK -> {args.out}")


if __name__ == "__main__":
    main()
