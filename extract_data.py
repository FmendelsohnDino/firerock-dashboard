#!/usr/bin/env python3
"""
FireRock Dashboard Data Extractor
Reads the daily Excel file and writes data.json for the dashboard
"""
import json, sys, os, zipfile
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    os.system('pip install openpyxl -q')
    from openpyxl import load_workbook


class ExtractError(Exception):
    """Raised when the input file or extracted data fails validation."""
    pass


def _check_real_xlsx(filepath):
    """A real .xlsx is a zip archive. Catch CSVs/HTML login pages renamed to .xlsx
    before openpyxl throws a cryptic BadZipFile."""
    with open(filepath, 'rb') as fh:
        head = fh.read(8)
    if head[:2] != b'PK':
        # Peek at the start to give a useful message (HTML login page, CSV, etc.)
        with open(filepath, 'rb') as fh:
            snippet = fh.read(200).decode('utf-8', 'replace').strip().replace('\n', ' ')
        raise ExtractError(
            f"'{filepath}' is not a real Excel file (no PK/zip signature). "
            f"It's likely a CSV or an HTML page saved with an .xlsx extension. "
            f"First bytes: {snippet[:120]!r}"
        )


def extract(filepath):
    _check_real_xlsx(filepath)
    try:
        wb = load_workbook(filepath, data_only=True)
    except zipfile.BadZipFile:
        raise ExtractError(
            f"'{filepath}' could not be opened as an Excel workbook. "
            f"Confirm a real .xlsx was uploaded, not a CSV or login page."
        )
    if 'DASHBOARD' not in wb.sheetnames:
        raise ExtractError(
            f"Workbook has no 'DASHBOARD' sheet. Found: {wb.sheetnames}"
        )
    ws = wb['DASHBOARD']
    wr = wb['REVENUE']

    def gc(sheet, r, c):
        v = sheet.cell(r, c).value
        return v if v is not None else 0

    # Date
    date_val = gc(ws, 3, 2)
    if hasattr(date_val, 'strftime'):
        report_date = date_val.strftime('%b %d, %Y')
    else:
        report_date = str(date_val) if date_val else datetime.now().strftime('%b %d, %Y')

    data = {
        "reportDate":    report_date,
        "workDaysMonth": gc(ws,5,4),
        "daysIn":        gc(ws,6,4),
        "pctMonth":      gc(ws,6,5),
        "workDaysYear":  gc(ws,5,11),
        "ytdDays":       gc(ws,6,11),
        "pctYear":       gc(ws,6,12),
        "mtdOrders":     gc(ws,30,3),
        "mtdPY":         gc(ws,30,4),
        "mtdQuota":      gc(ws,30,5),
        "mtdPctQ":       gc(ws,30,6),
        "mtdFore":       gc(ws,30,7),
        "forePct":       gc(ws,30,8),
        "ytdOrders":     gc(ws,30,14),
        "ytdPY":         gc(ws,30,15),
        "annualQuota":   gc(ws,30,19),   # FIXED: col 19 (S, "Annual Quota"), was 18 (R, "% Quota Trend")
        "backlog":       gc(ws,30,12),
        "revMTD":        gc(ws,45,10),
        "revBgtMTD":     gc(ws,45,11),
        "revPctMTD":     gc(ws,45,12),
        "revYTD":        gc(ws,45,16),
        "revBgtYTD":     gc(ws,45,17),
        "revPctYTD":     gc(ws,45,18),
        "reps": [],
        "prods": [],
        "revMonthly": [],
        "monthTotals": []
    }

    # Reps rows 14-23
    rep_names = ['ERIC BLASKOWSKI','KEVIN KEITH','THORNTON COLE','MICHAEL KUHLMAN',
                 'HAMP BRILEY','GRAYSON MOOREHEAD','BEAU BOUTHILLIER','SUZANNE HORST',
                 'TIFFANY HASKEKER','TOTAL FIREROCK SALES TEAM']
    for i, row in enumerate(range(14, 24)):
        data['reps'].append({
            "name":    gc(ws,row,2) or rep_names[i],
            "mtd":     gc(ws,row,3),
            "py":      gc(ws,row,4),
            "quota":   gc(ws,row,5),
            "pctQ":    gc(ws,row,6),
            "fore":    gc(ws,row,7),
            "forePct": gc(ws,row,8),
            "ytd":     gc(ws,row,14),
            "annQ":    gc(ws,row,19)
        })

    # Products rows 36-45
    prod_names = ['FIREPLACE/PP','PAVER/PP','ROOFING','FLOORING','INSTALLATION',
                  'SWD','OTHER','TOTAL','FREIGHT','TOTAL W/ FREIGHT']
    for i, row in enumerate(range(36, 46)):
        data['prods'].append({
            "name":    gc(ws,row,2) or prod_names[i],
            "mtd":     gc(ws,row,3),
            "quota":   gc(ws,row,5),
            "pctQ":    gc(ws,row,6),
            "revMTD":  gc(ws,row,10),
            "revBgt":  gc(ws,row,11),
            "revPct":  gc(ws,row,12),
            "backlog": gc(ws,row,14),
            "revYTD":  gc(ws,row,16)
        })

    # Revenue monthly rows 4-19
    for row in range(4, 20):
        name = gc(wr,row,1)
        if not name:
            continue
        # Skip the column-header row ("SALES REP | Jan | Feb ...") — it is not a person,
        # and its month cells are text, which the dashboard would render as "$NaN".
        if str(name).strip().upper() in ('SALES REP', 'REP'):
            continue

        def num(v):
            # Coerce blanks/text to 0 so a stray label can never become $NaN downstream.
            return v if isinstance(v, (int, float)) else 0

        data['revMonthly'].append({
            "name":  str(name),
            "jan":   num(gc(wr,row,2)),
            "feb":   num(gc(wr,row,3)),
            "mar":   num(gc(wr,row,4)),
            "apr":   num(gc(wr,row,5)),
            "may":   num(gc(wr,row,6)),
            "total": num(gc(wr,row,7))
        })

    # Month totals from grand total row
    gt_row = None
    for row in range(4, 25):
        if gc(wr,row,1) == 'Grand Total':
            gt_row = row
            break
    if gt_row:
        data['monthTotals'] = [
            {"mo":"Jan","v":gc(wr,gt_row,2)},
            {"mo":"Feb","v":gc(wr,gt_row,3)},
            {"mo":"Mar","v":gc(wr,gt_row,4)},
            {"mo":"Apr","v":gc(wr,gt_row,5)},
            {"mo":"May","v":gc(wr,gt_row,6)}
        ]

    data['lastUpdated'] = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

    _validate(data)
    return data


def _validate(data):
    """Sanity checks. Fail loudly instead of writing nonsense the dashboard
    will silently turn into a 615,000,000% tile."""
    errors = []

    aq = data['annualQuota']
    ytd = data['ytdOrders']

    # Annual quota must be a real dollar figure, not a percentage (the original bug).
    if not isinstance(aq, (int, float)):
        errors.append(f"annualQuota is not numeric: {aq!r}")
    elif aq < 1_000_000:
        errors.append(
            f"annualQuota = {aq:,.2f} is implausibly small (< $1M). "
            f"Likely reading the wrong column (a percentage), not 'Annual Quota'."
        )
    elif ytd and aq < ytd:
        errors.append(
            f"annualQuota ({aq:,.0f}) is less than YTD orders ({ytd:,.0f}) — impossible."
        )

    # A couple of cheap cross-checks on the other headline numbers.
    if data['mtdOrders'] and data['mtdOrders'] < 0:
        errors.append(f"mtdOrders is negative: {data['mtdOrders']}")
    if not (0 <= data['pctMonth'] <= 1.5):
        errors.append(f"pctMonth out of range: {data['pctMonth']}")

    if errors:
        raise ExtractError("Validation failed:\n  - " + "\n  - ".join(errors))


if __name__ == '__main__':
    filepath = sys.argv[1] if len(sys.argv) > 1 else 'sales_report.xlsx'
    if not os.path.exists(filepath):
        print(f'File not found: {filepath}')
        sys.exit(1)
    try:
        data = extract(filepath)
    except ExtractError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
    os.makedirs('data', exist_ok=True)
    with open('data/data.json', 'w') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"✓ Extracted {len(data['reps'])} reps, {len(data['prods'])} products")
    print(f"✓ MTD Orders: ${data['mtdOrders']:,.0f}")
    print(f"✓ Annual Quota: ${data['annualQuota']:,.0f}  (YTD {data['ytdOrders']/data['annualQuota']*100:.1f}% of annual)")
    print(f"✓ Written to data/data.json")
